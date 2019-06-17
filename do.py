from openpyxl import Workbook
wb=Workbook()
ws=wb.create_sheet("1")
f=open("1.txt")
k=1
for l in f.readlines():
    x=float(l)
    ws["A"+str(k)]=x
    k=k+1
wb.save("1.xlsx")